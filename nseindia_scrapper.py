# -*- coding: utf-8 -*-
debug = True
debug = False

#################### IMPORTS #####################################
from pprint import pprint
from fastnumbers import fast_real
from inspect import getmembers
from openpyxl import Workbook
from openpyxl import load_workbook
from lxml import html
from random import choice
from datetime import date
import queue
import requests
import urllib
import time
import sys
import requests.packages.urllib3
import csv
import os
import ast
import re
from PIL import Image
import pytesseract
import datetime
import json

args = sys.argv
if len(args) > 1:
	args[0] = args[1]
else:
	args = False

def var_dump(var):
    pprint(getmembers(var))


now = str(date.today())
pytesseract.pytesseract.tesseract_cmd = 'Tesseract-OCR/tesseract'
##################### GLOBAL VARS ####################################

desktop_agents = ['Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36',
                 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36',
                 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36',
                 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_1) AppleWebKit/602.2.14 (KHTML, like Gecko) Version/10.0.1 Safari/602.2.14',
                 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36',
                 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.98 Safari/537.36',
                 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.98 Safari/537.36',
                 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36',
                 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36',
                 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0']
 
def random_headers():
    return {'User-Agent': choice(desktop_agents),'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'}

######################## FUNCTIONS #################################
def dbug(msg):
	if debug:
		print(msg)

def dbug2(msg):
	if debug:
		print(msg)
		input()


def update_progress(progress, extra_text):
    barLength = 16 # Modify this to change the length of the progress bar
    status = ""
    if isinstance(progress, int):
        progress = float(progress)
    if not isinstance(progress, float):
        progress = 0
        status = "error: progress var must be float\r\n"
    if progress < 0:
        progress = 0
        status = "Wait...\r\n"
    if progress >= 1:
        progress = 1
        status = "COMPLETED.\r\n"
    block = int(round(barLength*progress))
    text = "\r{3} - Percentage: [{0}] {1}% {2}".format( "#"*block + "-"*(barLength-block), int(progress*100), status, extra_text)
    sys.stdout.write(text)
    sys.stdout.flush()

def get_tree(url):
	response = False
	while True:
		try:
			dbug(url)
			time.sleep(0)
			r = requests.get(url, headers=random_headers(), timeout=50)
			dbug(r)
			response = html.fromstring( r.content )
			break
		except:
			print("Problema de conexión, intentando de nuevo en 7 segundos...")
			time.sleep(7)
	return response

def request_page(url):
	response = False
	while True:
		try:
			dbug(url)
			time.sleep(0)
			r = requests.get(url, headers=random_headers(), timeout=50)
			dbug(r)
			response = r
			break
		except:
			print("Problema de conexión, intentando de nuevo en 7 segundos...")
			time.sleep(7)
	return response

def is_numeric(val):
    return str(val).replace('.','',1).isdigit()

def strarr_to_num(arr):
    return [int(float(x)) if int(float(x)) == float(x) else float(x) for x in arr]

def coerce(x):
    # it may be already int or float 
    if isinstance(x, (int, float)):
        return x
    # all int like strings can be converted to float so int tries first 
    try:
        return int(x)
    except (TypeError, ValueError):
        pass
    try:
        return float("{0:.2f}".format(x))
    except (TypeError, ValueError):
        return x

#rows = page.xpath('//table[@id="octable"]//tr') #get all rows
#cells = rows[2].xpath('./td/text()') #get cells. Starting from row[2]

def get_stock_data(symbol):
    page = get_tree("https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?symbol={}".format(symbol))
    rows = page.xpath('//table[@id="octable"]//tr') #get all rows

    data = [["-", "OI", "Chng in OI", "Volume", "IV", "LTP", "Net Chng", "BidQty", "BidPrice", "AskPrice", "AskQty", "Strike Price", "BidQty", "BidPrice", "AskPrice", "AskQty", "Net Chng", "LTP", "IV", "Volume", "Chng in OI", "OI", "-"]]
    rows = rows[2:]
    for r in rows:       
        row_data = [rr.text_content() for rr in r.xpath('./td')] #get row_data. Starting from row[2]
        row_data = [item.strip().replace(',', '') for item in row_data]
        row_data = [ coerce( iii ) for iii in row_data]
  
        data.append( row_data )    

    total_row = [''] * 23
    last_row = data[-1]
    total_row[0] = "Total"
    total_row[1] = last_row[1]
    total_row[3] = last_row[3]
    total_row[19] = last_row[5]
    total_row[21] = last_row[7]
    total_row[22] = "Total"

    data.pop()
    data.append( total_row )

    return data

def get_stocks(filepath):
    wb = load_workbook(filename = filepath)
    wb = wb.active
    stocks = []
    for i in range(2,999):
        stocks.append(wb["A{}".format(i)].value)
    return list(filter(None, stocks))

def scrape_data():
    stocks = get_stocks('multiple pages.xlsm')

    if not os.path.exists("scraped_data_{}".format(now)):
        os.makedirs("scraped_data_{}".format(now))

    i = 1;
    for stock in stocks:
        update_progress((float(i)/float(len(stocks))), ("STOCKS remaining {}/{}".format(i, len(stocks))))
        if stock.strip() != "":
            table_data = get_stock_data(stock)

            wb = Workbook()
            ws = wb.active
            for row in table_data:
                ws.append(row)
            
            data_saved = False
            while not data_saved:
                try:
                    wb.save("scraped_data_{}/{}.xlsx".format(now,stock))
                    data_saved = True
                    pass
                except Exception as e:
                    print("\nAn error ocurred: {}".format(e))
                    input("Make sure you don't have an excel file opened. Press Enter to retry")
                    pass
        i += 1;

print("##############################################")
print("#                                            #")
print("#            NseindiaScrapper v1.00          #")
print("#                                            #")
print("##############################################")
print("")

input("Press ENTER to start scrapping.")
scrape_data()
input("\nDONE. Scrapping has finished. Press ENTER to exit.")
exit()
